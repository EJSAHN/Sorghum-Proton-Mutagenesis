#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
95_spatial_validation_tables.py

Additional table-only analyses for spatial mutagenesis validation.

This script extends the main tables-only pipeline with analyses that quantify:
  1) gamma high-load samples versus proton samples,
  2) mutation density normalized by local GBS-callable site density,
  3) alternative high-density track definitions,
  4) dose-response regression parameters, and
  5) coding-region fraction models with radiation type as a covariate.

"""
from __future__ import annotations

import argparse
import gzip
import json
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats

try:
    from pipeline_utils import find_first_existing, resolve_project_paths
except ImportError:  # allow running from outside scripts/ if needed
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from pipeline_utils import find_first_existing, resolve_project_paths


CHR10 = [str(i) for i in range(1, 11)]
DEFAULT_WINDOW_SIZES = [250_000, 500_000, 1_000_000]
DEFAULT_PERCENTILES = [90.0, 95.0, 99.0]
DEFAULT_RANDOM_SEED = 42

CONTIG_MAP = {
    "CM027680": "1", "CM027681": "2", "CM027682": "3", "CM027683": "4",
    "CM027684": "5", "CM027685": "6", "CM027686": "7", "CM027687": "8",
    "CM027688": "9", "CM027689": "10",
    "CM027680.1": "1", "CM027681.1": "2", "CM027682.1": "3", "CM027683.1": "4",
    "CM027684.1": "5", "CM027685.1": "6", "CM027686.1": "7", "CM027687.1": "8",
    "CM027688.1": "9", "CM027689.1": "10",
    "Chr01": "1", "Chr02": "2", "Chr03": "3", "Chr04": "4", "Chr05": "5",
    "Chr06": "6", "Chr07": "7", "Chr08": "8", "Chr09": "9", "Chr10": "10",
    "chr01": "1", "chr02": "2", "chr03": "3", "chr04": "4", "chr05": "5",
    "chr06": "6", "chr07": "7", "chr08": "8", "chr09": "9", "chr10": "10",
    "1": "1", "2": "2", "3": "3", "4": "4", "5": "5",
    "6": "6", "7": "7", "8": "8", "9": "9", "10": "10",
}

CALLABLE_IUPAC = set("ACGTRYSWKM")
MISSING_IUPAC = {"N", "", "NA", "NAN", "NONE"}


@dataclass(frozen=True)
class RegressionResult:
    term: str
    estimate: float
    std_error: float
    t_stat: float
    p_value: float
    ci95_low: float
    ci95_high: float


def parse_int_list(text: str) -> List[int]:
    vals: List[int] = []
    for tok in str(text).split(","):
        tok = tok.strip().replace("_", "")
        if tok:
            vals.append(int(tok))
    return sorted(set(vals))


def parse_float_list(text: str) -> List[float]:
    vals: List[float] = []
    for tok in str(text).split(","):
        tok = tok.strip()
        if tok:
            vals.append(float(tok))
    return sorted(set(vals))


def normalize_chrom(name: object) -> str:
    s = str(name).strip()
    if not s:
        return ""
    if s in CONTIG_MAP:
        return CONTIG_MAP[s]
    s0 = re.split(r"[ \t|;]", s)[0]
    if s0 in CONTIG_MAP:
        return CONTIG_MAP[s0]
    base = s0.split(".")[0]
    if base in CONTIG_MAP:
        return CONTIG_MAP[base]
    s2 = s0.lower().replace("chromosome", "").replace("chr", "").strip()
    s2 = s2.replace("_", " ").replace("-", " ").strip()
    m = re.match(r"^0*([1-9]|10)\b", s2)
    if m:
        return str(int(m.group(1)))
    return s0


def normalize_sample_header(x: object) -> Optional[int]:
    if isinstance(x, (int, np.integer)):
        return int(x)
    s = str(x).strip()
    if s.isdigit():
        return int(s)
    m = re.fullmatch(r"(\d+)_\d+", s)
    if m:
        return int(m.group(1))
    return None


def is_callable_genotype(x: object) -> bool:
    if x is None:
        return False
    if isinstance(x, float) and math.isnan(x):
        return False
    s = str(x).strip().upper()
    if s in MISSING_IUPAC:
        return False
    return s in CALLABLE_IUPAC


def read_csv_any(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Required file not found: {path}")
    if str(path).endswith(".gz"):
        return pd.read_csv(path, compression="gzip", low_memory=False)
    return pd.read_csv(path, low_memory=False)


def gini(values: Iterable[float]) -> float:
    x = np.asarray(list(values), dtype=float)
    x = x[np.isfinite(x)]
    if x.size == 0:
        return float("nan")
    if np.any(x < 0):
        x = x - np.min(x)
    total = float(np.sum(x))
    if total <= 0:
        return 0.0
    x = np.sort(x)
    n = x.size
    idx = np.arange(1, n + 1, dtype=float)
    return float(np.sum((2.0 * idx - n - 1.0) * x) / (n * total))


def safe_float(x: object) -> float:
    try:
        return float(x)
    except Exception:
        return float("nan")


def cohen_d(x: Sequence[float], y: Sequence[float]) -> float:
    a = np.asarray(x, dtype=float)
    b = np.asarray(y, dtype=float)
    a = a[np.isfinite(a)]
    b = b[np.isfinite(b)]
    if a.size < 2 or b.size < 2:
        return float("nan")
    s1 = np.var(a, ddof=1)
    s2 = np.var(b, ddof=1)
    pooled = ((a.size - 1) * s1 + (b.size - 1) * s2) / (a.size + b.size - 2)
    if pooled <= 0:
        return float("nan")
    return float((np.mean(a) - np.mean(b)) / math.sqrt(pooled))


def cliffs_delta(x: Sequence[float], y: Sequence[float]) -> float:
    a = np.asarray(x, dtype=float)
    b = np.asarray(y, dtype=float)
    a = a[np.isfinite(a)]
    b = b[np.isfinite(b)]
    if a.size == 0 or b.size == 0:
        return float("nan")
    gt = 0
    lt = 0
    for v in a:
        gt += int(np.sum(v > b))
        lt += int(np.sum(v < b))
    return float((gt - lt) / (a.size * b.size))


def compare_two_groups(df: pd.DataFrame, group_col: str, group_a: str, group_b: str, metrics: Sequence[str]) -> pd.DataFrame:
    rows: List[dict] = []
    for metric in metrics:
        if metric not in df.columns:
            continue
        a = pd.to_numeric(df.loc[df[group_col] == group_a, metric], errors="coerce").dropna().to_numpy(float)
        b = pd.to_numeric(df.loc[df[group_col] == group_b, metric], errors="coerce").dropna().to_numpy(float)
        if a.size == 0 or b.size == 0:
            rows.append({
                "metric": metric, "group_a": group_a, "group_b": group_b,
                "n_a": int(a.size), "n_b": int(b.size),
            })
            continue
        t_p = float("nan")
        t_stat = float("nan")
        if a.size >= 2 and b.size >= 2:
            tres = stats.ttest_ind(a, b, equal_var=False, nan_policy="omit")
            t_stat = float(tres.statistic)
            t_p = float(tres.pvalue)
        mw_p = float("nan")
        mw_u = float("nan")
        if a.size >= 1 and b.size >= 1:
            try:
                mres = stats.mannwhitneyu(a, b, alternative="two-sided")
                mw_u = float(mres.statistic)
                mw_p = float(mres.pvalue)
            except ValueError:
                pass
        rows.append({
            "metric": metric,
            "group_a": group_a,
            "group_b": group_b,
            "n_a": int(a.size),
            "n_b": int(b.size),
            "mean_a": float(np.mean(a)),
            "mean_b": float(np.mean(b)),
            "median_a": float(np.median(a)),
            "median_b": float(np.median(b)),
            "sd_a": float(np.std(a, ddof=1)) if a.size > 1 else float("nan"),
            "sd_b": float(np.std(b, ddof=1)) if b.size > 1 else float("nan"),
            "difference_mean_a_minus_b": float(np.mean(a) - np.mean(b)),
            "welch_t": t_stat,
            "welch_p": t_p,
            "mann_whitney_u": mw_u,
            "mann_whitney_p": mw_p,
            "cohen_d_a_minus_b": cohen_d(a, b),
            "cliffs_delta_a_minus_b": cliffs_delta(a, b),
        })
    return pd.DataFrame(rows)


def read_events(events_path: Path) -> pd.DataFrame:
    events = read_csv_any(events_path)
    required = {"sample_id", "rad_type", "contig", "pos"}
    missing = required - set(events.columns)
    if missing:
        raise ValueError(f"Events file missing required columns: {sorted(missing)}")
    events = events.copy()
    events["sample_id"] = pd.to_numeric(events["sample_id"], errors="coerce").astype("Int64")
    events = events.dropna(subset=["sample_id"]).copy()
    events["sample_id"] = events["sample_id"].astype(int)
    events["rad_type"] = events["rad_type"].astype(str).str.lower().str.strip()
    if "dose_Gy" not in events.columns:
        events["dose_Gy"] = np.nan
    events["dose_Gy"] = pd.to_numeric(events["dose_Gy"], errors="coerce")
    events["chrom"] = events["contig"].apply(normalize_chrom)
    events["pos"] = pd.to_numeric(events["pos"], errors="coerce")
    events = events.dropna(subset=["pos"]).copy()
    events["pos"] = events["pos"].astype(int)
    events = events[events["chrom"].isin(CHR10)].copy()
    return events


def read_metadata(meta_path: Path, events: pd.DataFrame) -> pd.DataFrame:
    if meta_path.exists():
        meta = read_csv_any(meta_path)
        meta = meta.copy()
        if "sample_id" in meta.columns:
            meta["sample_id"] = pd.to_numeric(meta["sample_id"], errors="coerce").astype("Int64")
            meta = meta.dropna(subset=["sample_id"]).copy()
            meta["sample_id"] = meta["sample_id"].astype(int)
        else:
            raise ValueError(f"Metadata file missing sample_id column: {meta_path}")
        if "rad_type" in meta.columns:
            meta["rad_type"] = meta["rad_type"].astype(str).str.lower().str.strip()
        if "dose_Gy" in meta.columns:
            meta["dose_Gy"] = pd.to_numeric(meta["dose_Gy"], errors="coerce")
    else:
        meta = pd.DataFrame()

    event_meta = events.groupby("sample_id").agg(
        rad_type_from_events=("rad_type", "first"),
        dose_Gy_from_events=("dose_Gy", "first"),
        total_events_chr1_10=("pos", "count"),
    ).reset_index()

    if meta.empty:
        out = event_meta.rename(columns={"rad_type_from_events": "rad_type", "dose_Gy_from_events": "dose_Gy"})
    else:
        out = meta.merge(event_meta, on="sample_id", how="outer")
        if "rad_type" not in out.columns:
            out["rad_type"] = out["rad_type_from_events"]
        else:
            out["rad_type"] = out["rad_type"].fillna(out["rad_type_from_events"])
        if "dose_Gy" not in out.columns:
            out["dose_Gy"] = out["dose_Gy_from_events"]
        else:
            out["dose_Gy"] = out["dose_Gy"].fillna(out["dose_Gy_from_events"])
    out["rad_type"] = out["rad_type"].astype(str).str.lower().str.strip()
    return out


def event_burden_by_sample(events: pd.DataFrame) -> pd.DataFrame:
    return events.groupby(["sample_id", "rad_type", "dose_Gy"], dropna=False).size().reset_index(name="total_events_chr1_10")


def short_range_cluster_fraction(events: pd.DataFrame, cluster_bp: int) -> pd.DataFrame:
    ev = events.sort_values(["sample_id", "chrom", "pos"]).copy()
    ev["prev_pos"] = ev.groupby(["sample_id", "chrom"])["pos"].shift(1)
    ev["next_pos"] = ev.groupby(["sample_id", "chrom"])["pos"].shift(-1)
    ev["d_prev"] = ev["pos"] - ev["prev_pos"]
    ev["d_next"] = ev["next_pos"] - ev["pos"]
    ev["clustered_short_range"] = (ev["d_prev"] <= cluster_bp) | (ev["d_next"] <= cluster_bp)
    out = ev.groupby(["sample_id", "rad_type", "dose_Gy"], dropna=False).agg(
        total_events=("pos", "count"),
        clustered_events=("clustered_short_range", "sum"),
    ).reset_index()
    out["short_range_cluster_fraction"] = out["clustered_events"] / out["total_events"]
    out["cluster_bp"] = int(cluster_bp)
    return out


def event_window_counts(events: pd.DataFrame, window_size: int) -> pd.DataFrame:
    ev = events.copy()
    ev["window_size_bp"] = int(window_size)
    ev["window"] = (ev["pos"] // int(window_size)).astype(int)
    out = ev.groupby(["sample_id", "rad_type", "dose_Gy", "chrom", "window_size_bp", "window"], dropna=False).size().reset_index(name="event_count")
    return out


def collect_callable_windows(
    s4_path: Path,
    sample_ids: Sequence[int],
    window_sizes: Sequence[int],
    header_row: int,
    progress_every: int = 25_000,
) -> Dict[int, pd.DataFrame]:
    if not s4_path.exists():
        raise FileNotFoundError(f"Table S4 not found: {s4_path}")

    sample_set = set(int(x) for x in sample_ids)
    wb = load_workbook(s4_path, read_only=True, data_only=True)
    ws = wb.active

    header_excel_row = int(header_row) + 1
    data_start_row = header_excel_row + 1
    header = list(next(ws.iter_rows(min_row=header_excel_row, max_row=header_excel_row, values_only=True)))

    normalized_header = [str(x).strip().lower() if x is not None else "" for x in header]
    try:
        idx_contig = normalized_header.index("id")
    except ValueError:
        try:
            idx_contig = normalized_header.index("contig")
        except ValueError as e:
            raise ValueError("Table S4 header must contain 'id' or 'contig'.") from e
    try:
        idx_pos = normalized_header.index("pos")
    except ValueError as e:
        raise ValueError("Table S4 header must contain 'pos'.") from e

    sample_columns: List[Tuple[int, int]] = []
    seen_samples: set[int] = set()
    for idx, col in enumerate(header):
        sid = normalize_sample_header(col)
        if sid is not None and sid in sample_set and sid not in seen_samples:
            sample_columns.append((sid, idx))
            seen_samples.add(sid)

    if not sample_columns:
        wb.close()
        raise RuntimeError("No sample genotype columns from metadata were found in Table S4.")

    counts: Dict[int, Dict[Tuple[int, str, int], int]] = {int(w): defaultdict(int) for w in window_sizes}
    n_rows = 0
    n_chr_rows = 0

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        n_rows += 1
        if progress_every and n_rows % progress_every == 0:
            print(f"[INFO] Scanned {n_rows:,} Table S4 loci for callable-window denominators...")
        chrom = normalize_chrom(row[idx_contig] if idx_contig < len(row) else None)
        if chrom not in CHR10:
            continue
        pos_raw = row[idx_pos] if idx_pos < len(row) else None
        try:
            pos = int(pos_raw)
        except Exception:
            continue
        if pos <= 0:
            continue
        n_chr_rows += 1
        for sid, col_idx in sample_columns:
            call = row[col_idx] if col_idx < len(row) else None
            if not is_callable_genotype(call):
                continue
            for w in window_sizes:
                counts[int(w)][(sid, chrom, pos // int(w))] += 1

    wb.close()

    out: Dict[int, pd.DataFrame] = {}
    for w, d in counts.items():
        rows = [
            {"sample_id": sid, "chrom": chrom, "window_size_bp": int(w), "window": int(win), "callable_loci": int(v)}
            for (sid, chrom, win), v in d.items()
        ]
        out[int(w)] = pd.DataFrame(rows)
    print(f"[INFO] Callable-window scan complete: {n_rows:,} S4 rows scanned; {n_chr_rows:,} rows on chromosomes 1-10.")
    return out


def callable_normalized_density(events: pd.DataFrame, callable_by_window: Dict[int, pd.DataFrame], meta: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    per_window_parts: List[pd.DataFrame] = []
    for w, call_df in callable_by_window.items():
        ev_counts = event_window_counts(events, int(w))
        merged = call_df.merge(ev_counts, on=["sample_id", "chrom", "window_size_bp", "window"], how="left")
        merged["event_count"] = merged["event_count"].fillna(0).astype(int)
        sample_meta = meta[["sample_id", "rad_type", "dose_Gy"]].drop_duplicates("sample_id")
        merged = merged.merge(sample_meta, on="sample_id", how="left", suffixes=("", "_meta"))
        if "rad_type_meta" in merged.columns:
            merged["rad_type"] = merged["rad_type"].fillna(merged["rad_type_meta"])
            merged = merged.drop(columns=["rad_type_meta"])
        if "dose_Gy_meta" in merged.columns:
            merged["dose_Gy"] = merged["dose_Gy"].fillna(merged["dose_Gy_meta"])
            merged = merged.drop(columns=["dose_Gy_meta"])
        merged["rad_type"] = merged["rad_type"].astype(str).str.lower().str.strip()
        merged["normalized_events_per_1000_callable_loci"] = np.where(
            merged["callable_loci"] > 0,
            merged["event_count"] / merged["callable_loci"] * 1000.0,
            np.nan,
        )
        merged["window_start"] = merged["window"].astype(int) * int(w)
        merged["window_end"] = merged["window_start"] + int(w) - 1
        per_window_parts.append(merged)

    per_window = pd.concat(per_window_parts, ignore_index=True) if per_window_parts else pd.DataFrame()

    sample_rows: List[dict] = []
    group_cols = ["sample_id", "rad_type", "dose_Gy", "window_size_bp"]
    for keys, sub in per_window.groupby(group_cols, dropna=False):
        sid, rad, dose, w = keys
        event_count = pd.to_numeric(sub["event_count"], errors="coerce").fillna(0).to_numpy(float)
        norm = pd.to_numeric(sub["normalized_events_per_1000_callable_loci"], errors="coerce").fillna(0).to_numpy(float)
        callable_loci = pd.to_numeric(sub["callable_loci"], errors="coerce").fillna(0).to_numpy(float)
        sample_rows.append({
            "sample_id": int(sid),
            "rad_type": str(rad),
            "dose_Gy": safe_float(dose),
            "window_size_bp": int(w),
            "n_callable_windows": int(len(sub)),
            "callable_loci_total": int(np.sum(callable_loci)),
            "events_total_in_callable_windows": int(np.sum(event_count)),
            "windows_with_events": int(np.sum(event_count > 0)),
            "mean_normalized_events_per_1000_callable_loci": float(np.mean(norm)) if norm.size else float("nan"),
            "median_normalized_events_per_1000_callable_loci": float(np.median(norm)) if norm.size else float("nan"),
            "max_normalized_events_per_1000_callable_loci": float(np.max(norm)) if norm.size else float("nan"),
            "gini_window_event_counts": gini(event_count),
            "gini_normalized_window_density": gini(norm),
        })
    per_sample = pd.DataFrame(sample_rows)

    metric_cols = [
        "n_callable_windows", "callable_loci_total", "events_total_in_callable_windows", "windows_with_events",
        "mean_normalized_events_per_1000_callable_loci", "median_normalized_events_per_1000_callable_loci",
        "max_normalized_events_per_1000_callable_loci", "gini_window_event_counts", "gini_normalized_window_density",
    ]
    group_summary = per_sample.groupby(["window_size_bp", "rad_type"], dropna=False).agg(
        n_samples=("sample_id", "nunique"),
        **{f"{c}_mean": (c, "mean") for c in metric_cols},
        **{f"{c}_median": (c, "median") for c in metric_cols},
    ).reset_index()
    return per_window, per_sample, group_summary


def spatial_gini_callable_null(
    events: pd.DataFrame,
    callable_by_window: Dict[int, pd.DataFrame],
    meta: pd.DataFrame,
    window_size: int,
    permutations: int,
    random_seed: int,
) -> pd.DataFrame:
    rng = np.random.default_rng(int(random_seed))
    if window_size not in callable_by_window:
        return pd.DataFrame()
    call_df = callable_by_window[int(window_size)].copy()
    ev_counts = event_window_counts(events, int(window_size))
    merged = call_df.merge(ev_counts, on=["sample_id", "chrom", "window_size_bp", "window"], how="left")
    merged["event_count"] = merged["event_count"].fillna(0).astype(int)
    sample_meta = meta[["sample_id", "rad_type", "dose_Gy"]].drop_duplicates("sample_id")
    merged = merged.merge(sample_meta, on="sample_id", how="left", suffixes=("", "_meta"))
    if "rad_type_meta" in merged.columns:
        merged["rad_type"] = merged["rad_type"].fillna(merged["rad_type_meta"])
    if "dose_Gy_meta" in merged.columns:
        merged["dose_Gy"] = merged["dose_Gy"].fillna(merged["dose_Gy_meta"])

    rows: List[dict] = []
    for sid, sub in merged.groupby("sample_id"):
        callable_loci = pd.to_numeric(sub["callable_loci"], errors="coerce").fillna(0).to_numpy(float)
        obs_counts = pd.to_numeric(sub["event_count"], errors="coerce").fillna(0).to_numpy(float)
        total_callable = float(np.sum(callable_loci))
        total_events = int(np.sum(obs_counts))
        if total_callable <= 0 or total_events <= 0 or callable_loci.size == 0:
            continue
        probs = callable_loci / total_callable
        obs_gini = gini(obs_counts)
        sim_ginis = np.empty(int(permutations), dtype=float)
        for i in range(int(permutations)):
            sim_counts = rng.multinomial(total_events, probs)
            sim_ginis[i] = gini(sim_counts)
        null_mean = float(np.mean(sim_ginis))
        null_sd = float(np.std(sim_ginis, ddof=1)) if sim_ginis.size > 1 else float("nan")
        empirical_p_upper = float((np.sum(sim_ginis >= obs_gini) + 1) / (sim_ginis.size + 1))
        rows.append({
            "sample_id": int(sid),
            "rad_type": str(sub["rad_type"].iloc[0]).lower(),
            "dose_Gy": safe_float(sub["dose_Gy"].iloc[0]),
            "window_size_bp": int(window_size),
            "n_windows": int(len(sub)),
            "total_events": total_events,
            "total_callable_loci": int(total_callable),
            "observed_gini": obs_gini,
            "null_gini_mean": null_mean,
            "null_gini_sd": null_sd,
            "observed_minus_null_mean": obs_gini - null_mean,
            "observed_over_null_mean": obs_gini / null_mean if null_mean else float("nan"),
            "z_score_vs_null": (obs_gini - null_mean) / null_sd if null_sd and np.isfinite(null_sd) and null_sd > 0 else float("nan"),
            "empirical_p_observed_ge_null": empirical_p_upper,
            "permutations": int(permutations),
        })
    return pd.DataFrame(rows)


def track_sensitivity(events: pd.DataFrame, window_sizes: Sequence[int], percentiles: Sequence[float]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    per_sample_parts: List[pd.DataFrame] = []
    threshold_rows: List[dict] = []

    for w in window_sizes:
        counts = event_window_counts(events, int(w))
        nonzero = counts[counts["event_count"] > 0].copy()
        if nonzero.empty:
            continue
        for method in ["gamma_percentile", "pooled_percentile", "radiation_specific_percentile"]:
            if method == "gamma_percentile":
                source = nonzero.loc[nonzero["rad_type"] == "gamma", "event_count"].to_numpy(float)
                for p in percentiles:
                    if source.size == 0:
                        continue
                    threshold = float(np.percentile(source, p))
                    threshold_rows.append({
                        "window_size_bp": int(w), "threshold_method": method, "percentile": float(p),
                        "rad_type_for_threshold": "gamma", "threshold_count": threshold,
                        "n_nonzero_windows_for_threshold": int(source.size),
                    })
                    per_sample_parts.append(apply_track_threshold(counts, int(w), method, float(p), threshold, None))
            elif method == "pooled_percentile":
                source = nonzero["event_count"].to_numpy(float)
                for p in percentiles:
                    threshold = float(np.percentile(source, p))
                    threshold_rows.append({
                        "window_size_bp": int(w), "threshold_method": method, "percentile": float(p),
                        "rad_type_for_threshold": "all", "threshold_count": threshold,
                        "n_nonzero_windows_for_threshold": int(source.size),
                    })
                    per_sample_parts.append(apply_track_threshold(counts, int(w), method, float(p), threshold, None))
            else:
                for rad, sub_nonzero in nonzero.groupby("rad_type"):
                    source = sub_nonzero["event_count"].to_numpy(float)
                    for p in percentiles:
                        threshold = float(np.percentile(source, p))
                        threshold_rows.append({
                            "window_size_bp": int(w), "threshold_method": method, "percentile": float(p),
                            "rad_type_for_threshold": str(rad), "threshold_count": threshold,
                            "n_nonzero_windows_for_threshold": int(source.size),
                        })
                        per_sample_parts.append(apply_track_threshold(counts, int(w), method, float(p), threshold, str(rad)))

    per_sample = pd.concat(per_sample_parts, ignore_index=True) if per_sample_parts else pd.DataFrame()
    thresholds = pd.DataFrame(threshold_rows)

    if per_sample.empty:
        summary = pd.DataFrame()
    else:
        summary = per_sample.groupby(["window_size_bp", "threshold_method", "percentile", "rad_type"], dropna=False).agg(
            n_samples=("sample_id", "nunique"),
            threshold_count=("threshold_count", "first"),
            track_fraction_mean=("track_fraction", "mean"),
            track_fraction_median=("track_fraction", "median"),
            n_track_windows_mean=("n_track_windows", "mean"),
            track_events_mean=("track_events", "mean"),
            total_events_mean=("total_events", "mean"),
        ).reset_index()
    return per_sample, summary, thresholds


def apply_track_threshold(
    counts: pd.DataFrame,
    window_size: int,
    threshold_method: str,
    percentile: float,
    threshold: float,
    rad_specific: Optional[str],
) -> pd.DataFrame:
    if rad_specific is not None:
        target_counts = counts[counts["rad_type"] == rad_specific].copy()
    else:
        target_counts = counts.copy()
    rows: List[dict] = []
    for keys, sub in target_counts.groupby(["sample_id", "rad_type", "dose_Gy"], dropna=False):
        sid, rad, dose = keys
        total_events = int(sub["event_count"].sum())
        high = sub[sub["event_count"] >= threshold]
        track_events = int(high["event_count"].sum()) if not high.empty else 0
        rows.append({
            "sample_id": int(sid),
            "rad_type": str(rad),
            "dose_Gy": safe_float(dose),
            "window_size_bp": int(window_size),
            "threshold_method": threshold_method,
            "percentile": float(percentile),
            "threshold_count": float(threshold),
            "total_events": total_events,
            "track_events": track_events,
            "track_fraction": track_events / total_events if total_events > 0 else float("nan"),
            "n_track_windows": int(len(high)),
        })
    return pd.DataFrame(rows)


def high_load_gamma_comparison(
    events: pd.DataFrame,
    track_per_sample: pd.DataFrame,
    cluster_bp: int,
    default_window_size: int,
    default_percentile: float,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    burden = event_burden_by_sample(events).rename(columns={"total_events_chr1_10": "total_events"})
    short = short_range_cluster_fraction(events, cluster_bp)

    default_track = track_per_sample[
        (track_per_sample["window_size_bp"] == int(default_window_size))
        & (track_per_sample["threshold_method"] == "gamma_percentile")
        & (np.isclose(track_per_sample["percentile"].astype(float), float(default_percentile)))
    ].copy()
    keep_cols = ["sample_id", "track_fraction", "track_events", "n_track_windows", "threshold_count"]
    default_track = default_track[keep_cols].drop_duplicates("sample_id")

    metrics = burden.merge(default_track, on="sample_id", how="left")
    metrics = metrics.merge(short[["sample_id", "short_range_cluster_fraction", "clustered_events", "cluster_bp"]], on="sample_id", how="left")

    gamma = metrics[metrics["rad_type"] == "gamma"].copy()
    cutoff = float(gamma["total_events"].quantile(2.0 / 3.0)) if not gamma.empty else float("nan")
    metrics["comparison_group"] = metrics["rad_type"]
    metrics.loc[(metrics["rad_type"] == "gamma") & (metrics["total_events"] >= cutoff), "comparison_group"] = "gamma_high_load_tertile"
    metrics.loc[(metrics["rad_type"] == "gamma") & (metrics["total_events"] < cutoff), "comparison_group"] = "gamma_lower_two_tertiles"
    metrics.loc[metrics["rad_type"] == "proton", "comparison_group"] = "proton"
    metrics["gamma_high_load_cutoff_total_events"] = cutoff

    group_summary = metrics.groupby("comparison_group", dropna=False).agg(
        n_samples=("sample_id", "nunique"),
        total_events_mean=("total_events", "mean"),
        total_events_median=("total_events", "median"),
        track_fraction_mean=("track_fraction", "mean"),
        track_fraction_median=("track_fraction", "median"),
        short_range_cluster_fraction_mean=("short_range_cluster_fraction", "mean"),
        short_range_cluster_fraction_median=("short_range_cluster_fraction", "median"),
        n_track_windows_mean=("n_track_windows", "mean"),
    ).reset_index()

    comp_metrics = ["total_events", "track_fraction", "short_range_cluster_fraction", "n_track_windows", "track_events"]
    comparisons = pd.concat([
        compare_two_groups(metrics, "comparison_group", "proton", "gamma_high_load_tertile", comp_metrics),
        compare_two_groups(metrics, "comparison_group", "proton", "gamma_lower_two_tertiles", comp_metrics),
        compare_two_groups(metrics, "rad_type", "proton", "gamma", comp_metrics),
    ], ignore_index=True)
    comparisons.insert(0, "comparison_context", [
        "proton_vs_gamma_high_load_tertile" if i < len(comp_metrics) else
        "proton_vs_gamma_lower_two_tertiles" if i < len(comp_metrics) * 2 else
        "proton_vs_all_gamma"
        for i in range(len(comparisons))
    ])
    return metrics, group_summary, comparisons


def ols_fit(X: np.ndarray, y: np.ndarray, terms: Sequence[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    X = np.asarray(X, dtype=float)
    y = np.asarray(y, dtype=float)
    ok = np.isfinite(y) & np.all(np.isfinite(X), axis=1)
    X = X[ok]
    y = y[ok]
    n, p = X.shape
    if n <= p:
        coef = np.full(p, np.nan)
        rows = [{"term": t, "estimate": np.nan, "std_error": np.nan, "t_stat": np.nan, "p_value": np.nan, "ci95_low": np.nan, "ci95_high": np.nan} for t in terms]
        fit = pd.DataFrame([{"n": n, "df_residual": np.nan, "r_squared": np.nan, "rmse": np.nan}])
        return pd.DataFrame(rows), fit

    beta = np.linalg.lstsq(X, y, rcond=None)[0]
    yhat = X @ beta
    resid = y - yhat
    df = n - p
    rss = float(np.sum(resid ** 2))
    tss = float(np.sum((y - np.mean(y)) ** 2))
    mse = rss / df
    xtx_inv = np.linalg.inv(X.T @ X)
    se = np.sqrt(np.diag(xtx_inv) * mse)
    tvals = beta / se
    pvals = 2.0 * stats.t.sf(np.abs(tvals), df)
    tcrit = stats.t.ppf(0.975, df)
    rows = []
    for term, b, s, tv, pv in zip(terms, beta, se, tvals, pvals):
        rows.append({
            "term": term,
            "estimate": float(b),
            "std_error": float(s),
            "t_stat": float(tv),
            "p_value": float(pv),
            "ci95_low": float(b - tcrit * s),
            "ci95_high": float(b + tcrit * s),
        })
    fit = pd.DataFrame([{
        "n": int(n),
        "df_residual": int(df),
        "r_squared": float(1.0 - rss / tss) if tss > 0 else float("nan"),
        "rmse": float(math.sqrt(mse)),
        "rss": rss,
    }])
    return pd.DataFrame(rows), fit


def radclock_regression(per_mutant_path: Path, events: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if per_mutant_path.exists():
        df = read_csv_any(per_mutant_path)
        ycol = "induced_snvs" if "induced_snvs" in df.columns else None
        if ycol is None:
            df = event_burden_by_sample(events).rename(columns={"total_events_chr1_10": "induced_snvs"})
            ycol = "induced_snvs"
    else:
        df = event_burden_by_sample(events).rename(columns={"total_events_chr1_10": "induced_snvs"})
        ycol = "induced_snvs"

    df = df.copy()
    df["rad_type"] = df["rad_type"].astype(str).str.lower().str.strip()
    df["dose_Gy"] = pd.to_numeric(df["dose_Gy"], errors="coerce")
    df[ycol] = pd.to_numeric(df[ycol], errors="coerce")
    gamma = df[(df["rad_type"] == "gamma") & df["dose_Gy"].notna() & df[ycol].notna()].copy()
    x = gamma["dose_Gy"].to_numpy(float)
    y = gamma[ycol].to_numpy(float)
    X = np.column_stack([np.ones_like(x), x]) if x.size else np.empty((0, 2))
    coef, fit = ols_fit(X, y, ["intercept", "dose_Gy"])
    coef.insert(0, "model", "gamma_induced_events_vs_dose")

    prediction_rows: List[dict] = []
    if not coef.empty and coef["estimate"].notna().all():
        intercept = float(coef.loc[coef["term"] == "intercept", "estimate"].iloc[0])
        slope = float(coef.loc[coef["term"] == "dose_Gy", "estimate"].iloc[0])
        for dose in sorted(df["dose_Gy"].dropna().unique()):
            prediction_rows.append({"dose_Gy": float(dose), "gamma_regression_predicted_events": float(intercept + slope * float(dose))})
        proton = df[df["rad_type"] == "proton"].copy()
        if not proton.empty:
            proton_summary = proton.groupby("dose_Gy").agg(
                n_proton_samples=("sample_id", "nunique") if "sample_id" in proton.columns else (ycol, "count"),
                proton_mean_events=(ycol, "mean"),
                proton_median_events=(ycol, "median"),
                proton_sd_events=(ycol, "std"),
            ).reset_index()
            pred = pd.DataFrame(prediction_rows)
            proton_summary = proton_summary.merge(pred, on="dose_Gy", how="left")
            proton_summary["proton_mean_minus_gamma_prediction"] = proton_summary["proton_mean_events"] - proton_summary["gamma_regression_predicted_events"]
        else:
            proton_summary = pd.DataFrame()
    else:
        proton_summary = pd.DataFrame()

    return coef, fit, proton_summary


def cds_load_models(evolutionary_path: Path) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not evolutionary_path.exists():
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    df = read_csv_any(evolutionary_path)
    required = {"sample_id", "rad_type", "Total_Mutations", "CDS_Fraction"}
    missing = required - set(df.columns)
    if missing:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df = df.copy()
    df["rad_type"] = df["rad_type"].astype(str).str.lower().str.strip()
    df["Total_Mutations"] = pd.to_numeric(df["Total_Mutations"], errors="coerce")
    df["CDS_Fraction"] = pd.to_numeric(df["CDS_Fraction"], errors="coerce")
    df = df.dropna(subset=["Total_Mutations", "CDS_Fraction", "rad_type"]).copy()
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    load_mean = float(df["Total_Mutations"].mean())
    load_sd = float(df["Total_Mutations"].std(ddof=1))
    if not np.isfinite(load_sd) or load_sd <= 0:
        load_sd = 1.0
    df["load_z"] = (df["Total_Mutations"] - load_mean) / load_sd
    df["is_proton"] = (df["rad_type"] == "proton").astype(int)
    df["load_z_x_proton"] = df["load_z"] * df["is_proton"]

    X1 = np.column_stack([np.ones(len(df)), df["load_z"].to_numpy(float), df["is_proton"].to_numpy(float)])
    coef1, fit1 = ols_fit(X1, df["CDS_Fraction"].to_numpy(float), ["intercept", "load_z", "is_proton"])
    coef1.insert(0, "model", "CDS_Fraction ~ load_z + is_proton")
    fit1.insert(0, "model", "CDS_Fraction ~ load_z + is_proton")

    X2 = np.column_stack([np.ones(len(df)), df["load_z"].to_numpy(float), df["is_proton"].to_numpy(float), df["load_z_x_proton"].to_numpy(float)])
    coef2, fit2 = ols_fit(X2, df["CDS_Fraction"].to_numpy(float), ["intercept", "load_z", "is_proton", "load_z_x_proton"])
    coef2.insert(0, "model", "CDS_Fraction ~ load_z * is_proton")
    fit2.insert(0, "model", "CDS_Fraction ~ load_z * is_proton")

    coefs = pd.concat([coef1, coef2], ignore_index=True)
    fits = pd.concat([fit1, fit2], ignore_index=True)

    corr_rows: List[dict] = []
    for label, sub in [("all", df)] + [(str(rad), sub) for rad, sub in df.groupby("rad_type")]:
        x = sub["Total_Mutations"].to_numpy(float)
        y = sub["CDS_Fraction"].to_numpy(float)
        if len(sub) >= 3 and np.std(x) > 0 and np.std(y) > 0:
            pear = stats.pearsonr(x, y)
            spear = stats.spearmanr(x, y)
            corr_rows.append({
                "group": label,
                "n": int(len(sub)),
                "pearson_r": float(pear.statistic),
                "pearson_p": float(pear.pvalue),
                "spearman_rho": float(spear.statistic),
                "spearman_p": float(spear.pvalue),
                "load_min": float(np.min(x)),
                "load_max": float(np.max(x)),
                "cds_fraction_min": float(np.min(y)),
                "cds_fraction_max": float(np.max(y)),
            })
        else:
            corr_rows.append({"group": label, "n": int(len(sub))})
    correlations = pd.DataFrame(corr_rows)
    return coefs, fits, correlations


def config_sheet(args: argparse.Namespace, paths: object, derived: dict) -> pd.DataFrame:
    rows = [
        {"parameter": "project_root", "value": str(paths.root)},
        {"parameter": "events", "value": str(args.events)},
        {"parameter": "metadata", "value": str(args.meta)},
        {"parameter": "table_s4", "value": str(args.s4)},
        {"parameter": "window_sizes_bp", "value": ",".join(map(str, args.window_sizes))},
        {"parameter": "track_percentiles", "value": ",".join(map(str, args.track_percentiles))},
        {"parameter": "default_track_window_bp", "value": str(args.default_track_window)},
        {"parameter": "default_track_percentile", "value": str(args.default_track_percentile)},
        {"parameter": "cluster_bp", "value": str(args.cluster_bp)},
        {"parameter": "s4_header_row", "value": str(args.s4_header_row)},
        {"parameter": "spatial_null_window_bp", "value": str(args.spatial_null_window)},
        {"parameter": "spatial_null_permutations", "value": str(args.spatial_null_permutations)},
        {"parameter": "random_seed", "value": str(args.random_seed)},
        {"parameter": "output_workbook", "value": str(args.out)},
    ]
    for k, v in derived.items():
        rows.append({"parameter": str(k), "value": str(v)})
    return pd.DataFrame(rows)


def overview_sheet() -> pd.DataFrame:
    rows = [
        {"sheet": "analysis_overview", "description": "Short descriptions of workbook contents."},
        {"sheet": "configuration", "description": "Input paths and analysis parameters."},
        {"sheet": "event_burden", "description": "Per-sample induced SNV event counts on chromosomes 1-10."},
        {"sheet": "gamma_high_load_groups", "description": "Gamma high-load tertile assignment and proton comparison metrics."},
        {"sheet": "gamma_high_load_summary", "description": "Group summaries for proton, gamma high-load tertile, and lower-load gamma samples."},
        {"sheet": "gamma_high_load_tests", "description": "Welch and Mann-Whitney comparisons for clustering and track metrics."},
        {"sheet": "callable_norm_summary", "description": "Radiation-group summaries of mutation density normalized by local callable-locus density."},
        {"sheet": "callable_norm_per_sample", "description": "Per-sample callable-density-normalized window metrics."},
        {"sheet": "callable_norm_tests", "description": "Proton-vs-gamma tests for callable-density-normalized metrics."},
        {"sheet": "spatial_gini_null", "description": "Observed 1-Mb Gini values compared with callable-density-weighted permutation null expectations."},
        {"sheet": "track_thresholds", "description": "Thresholds used for alternative high-density track definitions."},
        {"sheet": "track_sensitivity_summary", "description": "Radiation-group track-fraction summaries across window sizes and threshold definitions."},
        {"sheet": "track_sensitivity_tests", "description": "Proton-vs-gamma tests across alternative track definitions."},
        {"sheet": "track_sensitivity_per_sample", "description": "Per-sample track metrics across alternative track definitions."},
        {"sheet": "radclock_coefficients", "description": "Gamma dose-response regression coefficients, standard errors, P-values, and 95% confidence intervals."},
        {"sheet": "radclock_fit", "description": "Gamma dose-response model fit statistics."},
        {"sheet": "proton_vs_gamma_prediction", "description": "Proton event burden at 300 Gy compared with gamma regression prediction."},
        {"sheet": "cds_load_coefficients", "description": "Linear models for CDS fraction as a function of mutation load and radiation type."},
        {"sheet": "cds_load_fit", "description": "Model fit summaries for CDS fraction models."},
        {"sheet": "cds_load_correlations", "description": "Group-specific correlations between mutation load and CDS fraction."},
        {"sheet": "callable_norm_windows", "description": "Per-sample, per-window event counts and callable-locus-normalized densities."},
    ]
    return pd.DataFrame(rows)


def write_workbook(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = name[:31]
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=safe_name, index=False)
            ws = writer.book[safe_name]
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def main() -> None:
    ap = argparse.ArgumentParser(description="Table-only spatial validation analyses for induced SNV events.")
    ap.add_argument("--project-root", default=None, help="Project root. Default: current working directory.")
    ap.add_argument("--events", default=None, help="Path to events_long.csv(.gz). Default: outputs_radclock/tables/events_long.csv.gz")
    ap.add_argument("--meta", default=None, help="Path to mutants metadata CSV. Default: outputs_radclock/tables/mutants_metadata.csv")
    ap.add_argument("--s4", default=None, help="Path to Table S4 Excel file. Default: inputs/Table S4.xlsx or Table_S4.xlsx")
    ap.add_argument("--per-mutant", default=None, help="Path to per_mutant_radclock.csv. Default: outputs_radclock/tables/per_mutant_radclock.csv")
    ap.add_argument("--evolutionary", default=None, help="Path to evolutionary_stats.csv. Default: outputs_functional_shielding/tables/evolutionary_stats.csv")
    ap.add_argument("--out", default=None, help="Output Excel workbook. Default: outputs_validation/Spatial_Validation_Tables.xlsx")
    ap.add_argument("--window-sizes", default="250000,500000,1000000", help="Comma-separated window sizes for track sensitivity.")
    ap.add_argument("--track-percentiles", default="90,95,99", help="Comma-separated percentiles for track thresholds.")
    ap.add_argument("--default-track-window", type=int, default=500_000, help="Default window size for high-load gamma comparison.")
    ap.add_argument("--default-track-percentile", type=float, default=95.0, help="Default gamma percentile for high-load gamma comparison.")
    ap.add_argument("--cluster-bp", type=int, default=10, help="Distance threshold for short-range cluster fraction.")
    ap.add_argument("--s4-header-row", type=int, default=2, help="0-based header row index in Table S4; default 2 means Excel row 3.")
    ap.add_argument("--spatial-null-window", type=int, default=1_000_000, help="Window size for callable-density-weighted Gini null.")
    ap.add_argument("--spatial-null-permutations", type=int, default=200, help="Number of permutations for callable-density-weighted Gini null.")
    ap.add_argument("--random-seed", type=int, default=DEFAULT_RANDOM_SEED, help="Random seed for permutation analyses.")
    args = ap.parse_args()

    args.window_sizes = parse_int_list(args.window_sizes)
    args.track_percentiles = parse_float_list(args.track_percentiles)
    if args.default_track_window not in args.window_sizes:
        args.window_sizes = sorted(set(args.window_sizes + [args.default_track_window]))
    if args.spatial_null_window not in args.window_sizes:
        args.window_sizes = sorted(set(args.window_sizes + [args.spatial_null_window]))

    paths = resolve_project_paths(args.project_root)
    events_path = Path(args.events) if args.events else (paths.outputs_radclock / "tables" / "events_long.csv.gz")
    meta_path = Path(args.meta) if args.meta else (paths.outputs_radclock / "tables" / "mutants_metadata.csv")
    per_mutant_path = Path(args.per_mutant) if args.per_mutant else (paths.outputs_radclock / "tables" / "per_mutant_radclock.csv")
    evolutionary_path = Path(args.evolutionary) if args.evolutionary else (paths.outputs_functional / "tables" / "evolutionary_stats.csv")
    s4_path = Path(args.s4) if args.s4 else find_first_existing([paths.inputs, paths.root], ["Table S4.xlsx", "Table_S4.xlsx"])
    out_path = Path(args.out) if args.out else (paths.root / "outputs_validation" / "Spatial_Validation_Tables.xlsx")

    args.events = str(events_path)
    args.meta = str(meta_path)
    args.s4 = str(s4_path)
    args.per_mutant = str(per_mutant_path)
    args.evolutionary = str(evolutionary_path)
    args.out = str(out_path)

    print(f"[INFO] Loading events: {events_path}")
    events = read_events(events_path)
    print(f"[INFO] Events retained on chromosomes 1-10: {len(events):,}")
    meta = read_metadata(meta_path, events)
    event_burden = event_burden_by_sample(events)
    sample_ids = sorted(event_burden["sample_id"].unique().tolist())

    print(f"[INFO] Scanning Table S4 callable windows: {s4_path}")
    callable_by_window = collect_callable_windows(s4_path, sample_ids, args.window_sizes, args.s4_header_row)

    print("[INFO] Computing callable-density-normalized mutation density...")
    callable_windows, callable_per_sample, callable_group_summary = callable_normalized_density(events, callable_by_window, meta)
    callable_tests_parts: List[pd.DataFrame] = []
    callable_metrics = [
        "mean_normalized_events_per_1000_callable_loci",
        "median_normalized_events_per_1000_callable_loci",
        "max_normalized_events_per_1000_callable_loci",
        "gini_window_event_counts",
        "gini_normalized_window_density",
        "windows_with_events",
    ]
    for w, sub in callable_per_sample.groupby("window_size_bp"):
        comp = compare_two_groups(sub, "rad_type", "proton", "gamma", callable_metrics)
        comp.insert(0, "window_size_bp", int(w))
        callable_tests_parts.append(comp)
    callable_tests = pd.concat(callable_tests_parts, ignore_index=True) if callable_tests_parts else pd.DataFrame()

    print("[INFO] Computing callable-density-weighted spatial Gini null...")
    spatial_null = spatial_gini_callable_null(
        events,
        callable_by_window,
        meta,
        args.spatial_null_window,
        args.spatial_null_permutations,
        args.random_seed,
    )

    print("[INFO] Computing track-definition sensitivity tables...")
    track_per_sample, track_summary, track_thresholds = track_sensitivity(events, args.window_sizes, args.track_percentiles)
    track_tests_parts: List[pd.DataFrame] = []
    for keys, sub in track_per_sample.groupby(["window_size_bp", "threshold_method", "percentile"], dropna=False):
        w, method, p = keys
        comp = compare_two_groups(sub, "rad_type", "proton", "gamma", ["track_fraction", "n_track_windows", "track_events", "total_events"])
        comp.insert(0, "window_size_bp", int(w))
        comp.insert(1, "threshold_method", str(method))
        comp.insert(2, "percentile", float(p))
        track_tests_parts.append(comp)
    track_tests = pd.concat(track_tests_parts, ignore_index=True) if track_tests_parts else pd.DataFrame()

    print("[INFO] Computing gamma high-load comparisons...")
    high_load_metrics, high_load_summary, high_load_tests = high_load_gamma_comparison(
        events,
        track_per_sample,
        args.cluster_bp,
        args.default_track_window,
        args.default_track_percentile,
    )

    print("[INFO] Computing RadClock regression parameters...")
    radclock_coef, radclock_fit, proton_prediction = radclock_regression(per_mutant_path, events)

    print("[INFO] Computing CDS fraction models...")
    cds_coef, cds_fit, cds_corr = cds_load_models(evolutionary_path)

    derived = {
        "n_events_chr1_10": len(events),
        "n_samples_with_events": event_burden["sample_id"].nunique(),
        "sample_ids_used_for_callable_scan": len(sample_ids),
    }

    sheets: Dict[str, pd.DataFrame] = {
        "analysis_overview": overview_sheet(),
        "configuration": config_sheet(args, paths, derived),
        "event_burden": event_burden,
        "gamma_high_load_groups": high_load_metrics,
        "gamma_high_load_summary": high_load_summary,
        "gamma_high_load_tests": high_load_tests,
        "callable_norm_summary": callable_group_summary,
        "callable_norm_per_sample": callable_per_sample,
        "callable_norm_tests": callable_tests,
        "spatial_gini_null": spatial_null,
        "track_thresholds": track_thresholds,
        "track_sensitivity_summary": track_summary,
        "track_sensitivity_tests": track_tests,
        "track_sensitivity_per_sample": track_per_sample,
        "radclock_coefficients": radclock_coef,
        "radclock_fit": radclock_fit,
        "proton_vs_gamma_prediction": proton_prediction,
        "cds_load_coefficients": cds_coef,
        "cds_load_fit": cds_fit,
        "cds_load_correlations": cds_corr,
        "callable_norm_windows": callable_windows,
    }

    print(f"[INFO] Writing workbook: {out_path}")
    write_workbook(out_path, sheets)
    print(f"[OK] Saved: {out_path}")


if __name__ == "__main__":
    main()
