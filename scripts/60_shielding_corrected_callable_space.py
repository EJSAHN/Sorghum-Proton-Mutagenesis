# -*- coding: utf-8 -*-
"""
60_shielding_corrected_callable_space.py

Functional shielding enrichment corrected by callable space bias (GBS ascertainment).

Method:
1) Observed counts: induced SNVs (events_annotated_v2.csv.gz) by radiation x functional category.
2) Expected probs: scan all callable loci in Table S4 (ref in A/C/G/T),
   map each locus (id,pos) to functional category using the SAME genomic_map logic as step 50:
      CDS > Intron(gene body) > Promoter > Intergenic
   and compute category frequencies = callable-space probabilities.
3) Expected counts: expected_prob[cat] * total_mutations(rad)
4) Significance: two-sided binomial test per category (k=observed, n=total_mutations, p=expected_prob)

Outputs:
  outputs_functional_shielding/tables/shielding_stats_corrected.csv
  outputs_functional_shielding/tables/expected_callable_space_probs.csv
"""

from __future__ import annotations

import argparse
import gzip
import re
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from scipy.stats import binomtest
from openpyxl import load_workbook
from tqdm import tqdm

from pipeline_utils import find_first_existing, resolve_project_paths


# CM contigs -> chr numbers
CONTIG_MAP = {
    "CM027680.1": "1",
    "CM027681.1": "2",
    "CM027682.1": "3",
    "CM027683.1": "4",
    "CM027684.1": "5",
    "CM027685.1": "6",
    "CM027686.1": "7",
    "CM027687.1": "8",
    "CM027688.1": "9",
    "CM027689.1": "10",
}
CHR10 = [str(i) for i in range(1, 11)]
PROMOTER_SIZE = 2000
CODE_TO_CAT = {0: "Intergenic", 1: "Promoter", 2: "Intron", 3: "CDS"}
CATEGORIES = ["Promoter", "CDS", "Intron", "Intergenic"]


def norm_chrom(x: object) -> str:
    """Normalize contig/seqid to chromosome number string '1'..'10' when possible."""
    s = str(x).strip()
    if not s:
        return ""
    if s in CONTIG_MAP:
        return CONTIG_MAP[s]
    m = re.search(r"(\d+)", s)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 10:
            return str(n)
    return s


def parse_gff_features(gff_path: Path) -> List[Tuple[str, str, int, int]]:
    """
    Parse GFF and keep only chr1-10 and features needed for genomic map.
    We keep exactly 'gene' and 'CDS' to match the proven v2 logic.
    Returns list of tuples: (chrom, ftype, start, end)
    """
    feats: List[Tuple[str, str, int, int]] = []
    with gzip.open(gff_path, "rt", encoding="utf-8", errors="replace") as f:
        for line in f:
            if not line or line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 9:
                continue

            seqid = parts[0]
            ftype = parts[2]
            start = parts[3]
            end = parts[4]

            chrom = norm_chrom(seqid)
            if chrom not in CHR10:
                continue
            if ftype not in ("gene", "CDS"):
                continue

            try:
                s = int(start)
                e = int(end)
            except ValueError:
                continue
            if s > e:
                s, e = e, s

            feats.append((chrom, ftype, s, e))
    return feats


def build_genomic_map(feats: List[Tuple[str, str, int, int]]) -> Dict[str, np.ndarray]:
    """
    Build chromosome arrays:
      0 Intergenic, 1 Promoter, 2 Intron(gene body), 3 CDS
    """
    by_chr: Dict[str, List[Tuple[str, int, int]]] = {}
    for chrom, ftype, s, e in feats:
        by_chr.setdefault(chrom, []).append((ftype, s, e))

    genomic_map: Dict[str, np.ndarray] = {}
    for chrom in sorted(by_chr.keys(), key=lambda x: int(x)):
        rows = by_chr[chrom]
        max_end = max(e for _t, _s, e in rows)
        arr = np.zeros(max_end + PROMOTER_SIZE + 10000 + 1, dtype=np.int8)

        # gene body + promoters
        for ftype, s, e in rows:
            if ftype == "gene":
                arr[s:e+1] = 2
                ps = max(1, s - PROMOTER_SIZE)
                pe = s - 1
                if pe >= ps:
                    region = arr[ps:pe+1]
                    arr[ps:pe+1] = np.where(region == 0, 1, region)

        # CDS (highest priority)
        for ftype, s, e in rows:
            if ftype == "CDS":
                arr[s:e+1] = 3

        genomic_map[chrom] = arr

    return genomic_map


def format_p_value(p: float) -> str:
    if p < 1e-100:
        return "<1e-100"
    return f"{p:.3e}"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--project-root", default=None)
    args = ap.parse_args()

    paths = resolve_project_paths(args.project_root)

    ann_path = paths.outputs_functional / "tables" / "events_annotated_v2.csv.gz"
    if not ann_path.exists():
        raise FileNotFoundError(f"Annotated events not found: {ann_path}. Run step 50 first.")

    s4_path = find_first_existing([paths.inputs, paths.root], ["Table S4.xlsx", "Table_S4.xlsx"])
    gff = find_first_existing(
        [paths.inputs, paths.root],
        [
            "SbicolorRio_468_v2.1.gene.gff3.gz",
            "SbicolorRio_468_v2.1.gene.gff3",
            "gene.gff3.gz",
            "genes.gff3.gz",
        ],
    )

    # -----------------------------
    # 1) Observed counts
    # -----------------------------
    print(f"[INFO] Loading annotation: {ann_path}")
    events = pd.read_csv(ann_path, compression="gzip", low_memory=False)

    # Observed counts (induced mutations) from Category assigned in step 50
    obs_counts = events.groupby(["rad_type", "Category"]).size().unstack(fill_value=0)
    for c in CATEGORIES:
        if c not in obs_counts.columns:
            obs_counts[c] = 0
    obs_counts = obs_counts[CATEGORIES].copy()

    print("[INFO] Observed mutation counts by category:")
    print(obs_counts)

    # -----------------------------
    # 2) Build genomic map (same logic as step 50)
    # -----------------------------
    print(f"[INFO] Parsing GFF3 for callable-space mapping (array): {gff}")
    feats = parse_gff_features(Path(gff))
    if not feats:
        raise RuntimeError("No gene/CDS features parsed from GFF on chr1-10. Check GFF feature types/seqids.")
    genomic_map = build_genomic_map(feats)

    # -----------------------------
    # 3) Expected probs from callable space (Table S4)
    # -----------------------------
    print(f"[INFO] Scanning Table S4 for callable loci: {s4_path}")
    wb = load_workbook(s4_path, read_only=True, data_only=True)
    ws = wb.active

    header = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))

    def hidx(name: str) -> int:
        try:
            return header.index(name)
        except ValueError as e:
            raise ValueError(f"Column '{name}' not found in Table S4 header.") from e

    IDX_CONTIG = hidx("id")
    IDX_POS = hidx("pos")
    IDX_REF = hidx("ref")

    bg_counts = {c: 0 for c in CATEGORIES}
    total_bg = 0

    row_iter = ws.iter_rows(min_row=4, values_only=True)
    it = tqdm(row_iter, desc="Callable-space scan", unit="loci")

    for row in it:
        ref = row[IDX_REF]
        if ref not in ("A", "C", "G", "T"):
            continue

        chrom = norm_chrom(row[IDX_CONTIG])
        if chrom not in genomic_map:
            continue

        pos = row[IDX_POS]
        if pos is None:
            continue
        try:
            pos_i = int(pos)
        except Exception:
            continue

        arr = genomic_map[chrom]
        if pos_i <= 0 or pos_i >= len(arr):
            continue

        cat = CODE_TO_CAT[int(arr[pos_i])]
        bg_counts[cat] += 1
        total_bg += 1

    wb.close()

    if total_bg == 0:
        raise RuntimeError(
            "No callable loci mapped to chr1-10 genomic arrays. "
            "Check Table S4 id/pos columns and coordinate system."
        )

    expected_probs = {c: (bg_counts[c] / total_bg) for c in CATEGORIES}
    print("[INFO] Expected callable-space probabilities:")
    for c in CATEGORIES:
        print(f"  {c:10s}: {expected_probs[c]:.6f}  (count={bg_counts[c]:,})")

    # Save expected probs
    probs_df = pd.DataFrame([{
        "Category": c,
        "callable_count": int(bg_counts[c]),
        "callable_prob": float(expected_probs[c]),
    } for c in CATEGORIES])
    out_probs = paths.outputs_functional / "tables" / "expected_callable_space_probs.csv"
    out_probs.parent.mkdir(parents=True, exist_ok=True)
    probs_df.to_csv(out_probs, index=False)
    print(f"[OK] Saved: {out_probs}")

    # -----------------------------
    # 4) Corrected shielding table
    # -----------------------------
    rows = []
    for rad in obs_counts.index:
        total_mut = int(obs_counts.loc[rad].sum())
        for cat in CATEGORIES:
            observed = int(obs_counts.loc[rad, cat])
            expected = expected_probs[cat] * total_mut
            enrichment = (observed / expected) if expected > 0 else float("nan")
            pval = binomtest(k=observed, n=total_mut, p=expected_probs[cat], alternative="two-sided").pvalue

            rows.append({
                "Radiation": rad,
                "Feature": cat,
                "Observed": observed,
                "Expected": float(expected),
                "Enrichment": float(enrichment),
                "P_value": format_p_value(float(pval)),
            })

    out_df = pd.DataFrame(rows)
    out_csv = paths.outputs_functional / "tables" / "shielding_stats_corrected.csv"
    out_df.to_csv(out_csv, index=False)
    print(f"[OK] Saved: {out_csv}")


if __name__ == "__main__":
    main()
