# -*- coding: utf-8 -*-
"""
10_event_burden_and_signatures.py

Event burden and mutational signatures
--------------------------------------
Goal: derive *induced* mutational signatures by comparing each mutant to its own non-irradiated parent (Origin line).

Tables-only workflow.

Inputs (inputs/):
  - Table S2.xlsx (or Table_S2.xlsx)
  - Table S4.xlsx (or Table_S4.xlsx)

Outputs (outputs_event_burden/):
  - tables/per_sample_event_burden.csv
  - tables/sub6_fractions.csv
  - tables/sub96_matrix.csv
  - tables/events_long.csv.gz
  - tables/mutants_metadata.csv
  - tables/analysis_config.json

Run (from project root):
  python scripts/10_event_burden_and_signatures.py --no-figures
"""
from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from pipeline_utils import find_first_existing, resolve_project_paths

# -------------------
# CONFIG
# -------------------
RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)

IUPAC: Dict[str, set] = {
    "A": {"A"}, "C": {"C"}, "G": {"G"}, "T": {"T"},
    "R": {"A", "G"}, "Y": {"C", "T"}, "S": {"G", "C"}, "W": {"A", "T"},
    "K": {"G", "T"}, "M": {"A", "C"},
    "B": {"C", "G", "T"}, "D": {"A", "G", "T"}, "H": {"A", "C", "T"}, "V": {"A", "C", "G"},
    "N": {"A", "C", "G", "T"},
    "-": set(), ".": set(),
}

COMP = {"A": "T", "T": "A", "C": "G", "G": "C"}

SUB6_ORDER = ["C>A", "C>G", "C>T", "T>A", "T>C", "T>G"]


def to_set(call) -> set:
    if call is None:
        return set()
    s = str(call).strip()
    if not s or s.lower() == "nan":
        return set()
    return IUPAC.get(s, set())


def is_unambig(call) -> bool:
    return len(to_set(call)) == 1


def get_base(call) -> Optional[str]:
    s = to_set(call)
    if len(s) != 1:
        return None
    return next(iter(s))


def revcomp(seq: str) -> str:
    seq = seq.upper()
    return "".join(COMP.get(b, "N") for b in reversed(seq))


def parse_treatment(t: object) -> Tuple[Optional[str], Optional[int]]:
    if t is None or (isinstance(t, float) and math.isnan(t)):
        return None, None
    s = str(t)
    m = re.search(r"(\d+)\s*Gy", s)
    dose = int(m.group(1)) if m else None
    if "Gamma" in s or "gamma" in s:
        return "gamma", dose
    if "Proton" in s or "proton" in s:
        return "proton", dose
    return "other", dose


def extract_trinuc(flank: object) -> Optional[str]:
    """Flanking_seq.(600bp) looks like ...CTC[G]GTT... ; return 3-mer centered on REF."""
    if flank is None:
        return None
    s = str(flank)
    i = s.find("[")
    j = s.find("]")
    if i < 1 or j != i + 2:
        return None
    prev_b = s[i - 1].upper()
    ref_b = s[i + 1].upper()
    next_b = s[i + 3].upper() if i + 3 < len(s) else None
    if next_b is None:
        return None
    tri = prev_b + ref_b + next_b
    if any(b not in "ACGT" for b in tri):
        return None
    return tri


def canon_sub_and_tri(ref: str, alt: str, tri: str) -> Tuple[str, str]:
    """
    Return canonical substitution (pyrimidine-centered) and oriented 3-mer.
    Standard: represent substitutions as C or T on reference Watson-Crick pair.
    """
    ref = ref.upper()
    alt = alt.upper()
    tri = tri.upper()
    if ref in ("A", "G"):
        ref = COMP[ref]
        alt = COMP[alt]
        tri = revcomp(tri)
    sub = f"{ref}>{alt}"
    return sub, tri


def sub96_key(ref: str, alt: str, tri: str) -> str:
    sub, tri2 = canon_sub_and_tri(ref, alt, tri)
    return f"{tri2[0]}[{sub}]{tri2[2]}"


def build_s2_table(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path, header=None)
    header = raw.iloc[1].tolist()
    df = raw.iloc[2:].copy()
    df.columns = header
    df = df.dropna(how="all").reset_index(drop=True)
    df["sample_id"] = np.arange(1, len(df) + 1, dtype=int)
    df["rad_type"], df["dose_Gy"] = zip(*df["Treatment \n(part)"].apply(parse_treatment))
    return df


@dataclass
class MutRec:
    line: str
    origin: str
    rad_type: str
    dose_Gy: int
    sample_id: int
    parent_id: int


class EventBurdenAndSignaturesPipeline:
    def __init__(self, table_s2: Path, table_s4: Path, out_dir: Path):
        self.table_s2 = table_s2
        self.table_s4 = table_s4
        self.out_dir = out_dir
        (self.out_dir / "tables").mkdir(parents=True, exist_ok=True)
        (self.out_dir / "cache").mkdir(parents=True, exist_ok=True)

    def load_metadata(self) -> List[MutRec]:
        s2 = build_s2_table(self.table_s2)

        # parent lookup (case-insensitive)
        line_to_sid = {str(l): int(sid) for l, sid in zip(s2["Lines"].astype(str), s2["sample_id"])}
        lower_to_sid = {k.lower(): v for k, v in line_to_sid.items()}

        parents = s2[s2["Types"] != "Mutant line"].copy()
        parent_lines_lower = set(parents["Lines"].astype(str).str.lower())
        mutants = s2[s2["Types"] == "Mutant line"].copy()

        mut_recs: List[MutRec] = []
        for r in mutants.itertuples(index=False):
            origin = str(getattr(r, "Origin"))
            parent_id = None
            if origin.lower() in parent_lines_lower:
                parent_id = lower_to_sid.get(origin.lower())
            if parent_id is None:
                continue
            mut_recs.append(
                MutRec(
                    line=str(getattr(r, "Lines")),
                    origin=origin,
                    rad_type=str(getattr(r, "rad_type")),
                    dose_Gy=int(getattr(r, "dose_Gy")),
                    sample_id=int(getattr(r, "sample_id")),
                    parent_id=int(parent_id),
                )
            )

        meta_df = pd.DataFrame([{
            "line": m.line,
            "origin": m.origin,
            "rad_type": m.rad_type,
            "dose_Gy": m.dose_Gy,
            "sample_id": m.sample_id,
            "parent_id": m.parent_id,
        } for m in mut_recs])
        meta_df.to_csv(self.out_dir / "tables" / "mutants_metadata.csv", index=False)

        return mut_recs

    def scan_s4_induced(self, mut_recs: List[MutRec], max_rows: Optional[int] = None):
        """
        Scan Table S4 in streaming mode; for each mutant, count induced SNVs at loci where:
          - parent base is unambiguous A/C/G/T
          - parent base equals REF (ancestral control)
          - mutant base is unambiguous and differs from parent/REF
          - trinucleotide context is parsable from Flanking_seq.(600bp)
        """
        from openpyxl import load_workbook
        from tqdm import tqdm

        wb = load_workbook(self.table_s4, read_only=True, data_only=True)
        ws = wb.active

        header = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))

        def hidx(name: str) -> int:
            try:
                return header.index(name)
            except ValueError as e:
                raise ValueError(f"Column '{name}' not found in Table S4 header.") from e

        IDX_CONTIG = hidx("id")
        IDX_POS = hidx("pos")
        IDX_REF = hidx("ref")
        IDX_FLANK = hidx("Flanking_seq.(600bp)")

        sid_to_col = {int(v): i for i, v in enumerate(header) if isinstance(v, int) and 1 <= int(v) <= 96}

        pairs: List[Tuple[MutRec, int, int]] = []
        for m in mut_recs:
            if m.sample_id not in sid_to_col or m.parent_id not in sid_to_col:
                continue
            pairs.append((m, sid_to_col[m.sample_id], sid_to_col[m.parent_id]))

        assessed = defaultdict(int)
        induced = defaultdict(int)
        sub6 = defaultdict(Counter)
        sub96 = defaultdict(Counter)
        events = []

        row_iter = ws.iter_rows(min_row=4, values_only=True)
        it = tqdm(row_iter, desc="Scanning S4 (induced SNVs vs parent)", unit="loci")

        for k, row in enumerate(it, start=1):
            if max_rows is not None and k > max_rows:
                break

            contig = row[IDX_CONTIG]
            pos = row[IDX_POS]
            ref = row[IDX_REF]

            if ref not in ("A", "C", "G", "T"):
                continue

            tri = extract_trinuc(row[IDX_FLANK])
            if tri is None:
                continue

            for m, c_mut, c_par in pairs:
                p_call = row[c_par]
                m_call = row[c_mut]
                if not is_unambig(p_call) or not is_unambig(m_call):
                    continue

                p_base = get_base(p_call)
                x_base = get_base(m_call)
                if p_base is None or x_base is None:
                    continue

                if p_base != ref:
                    continue

                assessed[m.sample_id] += 1
                if x_base == p_base:
                    continue

                induced[m.sample_id] += 1
                sub = canon_sub_and_tri(p_base, x_base, tri)[0]
                key96 = sub96_key(p_base, x_base, tri)
                sub6[m.sample_id][sub] += 1
                sub96[m.sample_id][key96] += 1

                events.append({
                    "line": m.line,
                    "origin": m.origin,
                    "rad_type": m.rad_type,
                    "dose_Gy": m.dose_Gy,
                    "sample_id": m.sample_id,
                    "parent_id": m.parent_id,
                    "contig": contig,
                    "pos": int(pos) if pos is not None else None,
                    "ref": p_base,
                    "alt": x_base,
                    "tri": tri,
                    "sub6": sub,
                    "sub96": key96,
                })

        wb.close()

        # per-mutant summary
        rows = []
        for m in mut_recs:
            sid = m.sample_id
            a = assessed.get(sid, 0)
            n = induced.get(sid, 0)
            rate = (n / a) if a > 0 else float("nan")
            c = sub6.get(sid, Counter())
            ts = c.get("C>T", 0) + c.get("T>C", 0)
            tv = (n - ts) if n >= ts else float("nan")
            ts_tv = (ts / tv) if (isinstance(tv, (int, float)) and tv not in (0, float("nan"))) else float("nan")
            rows.append({
                "line": m.line,
                "origin": m.origin,
                "rad_type": m.rad_type,
                "dose_Gy": m.dose_Gy,
                "sample_id": sid,
                "parent_id": m.parent_id,
                "assessed_loci": a,
                "induced_snvs": n,
                "induced_rate": rate,
                "ts_tv": ts_tv,
            })

        per_df = pd.DataFrame(rows)
        per_df.to_csv(self.out_dir / "tables" / "per_sample_event_burden.csv", index=False)

        # sub6 fractions
        sub6_rows = []
        for m in mut_recs:
            sid = m.sample_id
            total = sum(sub6[sid].values())
            for s in SUB6_ORDER:
                sub6_rows.append({
                    "line": m.line, "rad_type": m.rad_type, "dose_Gy": m.dose_Gy,
                    "sub6": s, "count": sub6[sid].get(s, 0),
                    "fraction": (sub6[sid].get(s, 0) / total) if total > 0 else float("nan"),
                })
        sub6_df = pd.DataFrame(sub6_rows)
        sub6_df.to_csv(self.out_dir / "tables" / "sub6_fractions.csv", index=False)

        # 96 matrix (wide, fractions)
        all96 = sorted({k for sid in sub96 for k in sub96[sid].keys()})
        mat = []
        for m in mut_recs:
            sid = m.sample_id
            vec = [sub96[sid].get(k, 0) for k in all96]
            total = sum(vec)
            frac = [(v / total) if total > 0 else 0.0 for v in vec]
            mat.append({
                "line": m.line, "rad_type": m.rad_type, "dose_Gy": m.dose_Gy,
                **{k: f for k, f in zip(all96, frac)}
            })
        sub96_wide = pd.DataFrame(mat)
        sub96_wide.to_csv(self.out_dir / "tables" / "sub96_matrix.csv", index=False)

        # events long
        ev_df = pd.DataFrame(events)
        ev_df.to_csv(self.out_dir / "tables" / "events_long.csv.gz", index=False, compression="gzip")

        meta = {
            "project": "Sorghum proton mutagenesis",
            "random_seed": RANDOM_SEED,
            "definition": "induced SNV where parent_call==REF and mutant_call!=REF (both unambiguous) with parsable trinuc context",
            "n_mutants_used": int(per_df.shape[0]),
            "table_s2": str(self.table_s2),
            "table_s4": str(self.table_s4),
        }
        (self.out_dir / "tables" / "analysis_config.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

        return per_df, sub6_df, sub96_wide

    def run(self, max_rows: Optional[int] = None) -> None:
        mut_recs = self.load_metadata()
        self.scan_s4_induced(mut_recs, max_rows=max_rows)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--project-root", default=None, help="Project root (default: CWD).")
    ap.add_argument("--max-rows", default=None, type=int, help="DEBUG: limit scanned loci in Table S4.")
    ap.add_argument("--no-figures", action="store_true", help="Accepted for compatibility; the workflow is table-only.")
    args = ap.parse_args()

    paths = resolve_project_paths(args.project_root)

    table_s2 = find_first_existing([paths.inputs, paths.root], ["Table S2.xlsx", "Table_S2.xlsx"])
    table_s4 = find_first_existing([paths.inputs, paths.root], ["Table S4.xlsx", "Table_S4.xlsx"])

    rc = EventBurdenAndSignaturesPipeline(table_s2=table_s2, table_s4=table_s4, out_dir=paths.outputs_event_burden)
    rc.run(max_rows=args.max_rows)


if __name__ == "__main__":
    main()
